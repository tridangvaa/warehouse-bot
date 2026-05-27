"""
Telegram Warehouse Management Bot
Mirrors SLT.xlsm structure (DMHANGHOA + GHISO) in Google Sheets.
Reads PDF/image phiếu nhập/xuất kho, appends to GHISO, calculates live stock.
"""

import os
import io
import json
import base64
import logging
import re
import time
from datetime import datetime
from dotenv import load_dotenv
from telegram import Update
from telegram.ext import Application, MessageHandler, CommandHandler, filters, ContextTypes
import anthropic
import gspread
import openpyxl
from google.oauth2.service_account import Credentials

load_dotenv()

logging.basicConfig(
    format="%(asctime)s | %(levelname)s | %(name)s | %(message)s",
    level=logging.INFO,
)
logger = logging.getLogger(__name__)

TELEGRAM_BOT_TOKEN      = os.environ["TELEGRAM_BOT_TOKEN"]
ANTHROPIC_API_KEY       = os.environ["ANTHROPIC_API_KEY"]
GOOGLE_SHEET_ID         = os.environ["GOOGLE_SHEET_ID"]
GOOGLE_CREDENTIALS_FILE = os.environ.get("GOOGLE_CREDENTIALS_FILE", "credentials.json")
TARGET_GROUP_ID         = os.environ.get("TARGET_GROUP_ID")
GOOGLE_SHEET_NAME       = os.environ.get("GOOGLE_SHEET_NAME", "DMHANGHOA")

# Single sheet for all data
SHEET_DMHANGHOA = GOOGLE_SHEET_NAME   # master item list + opening balance + min stock

DMHANGHOA_HEADERS = [
    "STT", "Kho", "Mã hàng", "Tên Hàng", "Đơn Vị Tính",
    "Tồn mức MIN",    # F (col 6)
    "Tồn đầu kỳ",    # G (col 7)
    "Nhập kỳ",        # H (col 8)
    "Xuất kỳ",        # I (col 9)
    "Tồn cuối",       # J (col 10) ← source of truth for current stock
]

# ── Stock warning levels (% of Tồn mức MIN) ───────────────────────────────────
# current_qty / min_qty ratio → warning level
WARN_L1 = 0.50   # 🟢 ≤ 50% of min
WARN_L2 = 0.30   # 🟠 ≤ 30% of min
WARN_L3 = 0.20   # 🔴 ≤ 20% of min


def stock_warning(current_qty: float, min_qty: float):
    """Return warning string if below threshold, else None."""
    if min_qty <= 0:
        return None
    ratio = current_qty / min_qty
    if ratio <= WARN_L3:
        return f"🔴 Mức 3 — Tồn {current_qty:.0f} ({ratio*100:.0f}% MIN) — CẦN NHẬP NGAY"
    if ratio <= WARN_L2:
        return f"🟠 Mức 2 — Tồn {current_qty:.0f} ({ratio*100:.0f}% MIN) — Sắp hết hàng"
    if ratio <= WARN_L1:
        return f"🟢 Mức 1 — Tồn {current_qty:.0f} ({ratio*100:.0f}% MIN) — Đang giảm"
    return None


claude = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)

_gc = None


def _get_gc():
    global _gc
    if _gc is None:
        creds_json = os.environ.get("GOOGLE_CREDENTIALS_JSON")
        if creds_json:
            info = json.loads(creds_json)
            creds = Credentials.from_service_account_info(
                info, scopes=["https://www.googleapis.com/auth/spreadsheets"]
            )
        else:
            creds = Credentials.from_service_account_file(
                GOOGLE_CREDENTIALS_FILE,
                scopes=["https://www.googleapis.com/auth/spreadsheets"],
            )
        _gc = gspread.authorize(creds)
    return _gc


def _ws() -> gspread.Worksheet:
    spreadsheet = _get_gc().open_by_key(GOOGLE_SHEET_ID)
    try:
        return spreadsheet.worksheet(SHEET_DMHANGHOA)
    except gspread.WorksheetNotFound:
        ws = spreadsheet.add_worksheet(title=SHEET_DMHANGHOA, rows=2000, cols=len(DMHANGHOA_HEADERS) + 2)
        return ws


def ensure_sheets():
    _ws()  # just verify the sheet exists


# ── DMHANGHOA helpers ──────────────────────────────────────────────────────────

def get_items() -> dict[str, dict]:
    """
    Read Bao_Cao_Ton_Kho layout:
      A(0)=empty, B(1)=STT, C(2)=Kho, D(3)=Mã hàng, E(4)=Tên, F(5)=ĐVT,
      G(6)=Tồn đầu, H(7)=Nhập kho, I(8)=Xuất bán, J(9)=Tồn cuối
    Data rows identified by STT (col B) being a positive integer.
    """
    ws = _ws()
    rows = ws.get_all_values()
    items = {}
    for i, row in enumerate(rows, start=1):
        if len(row) < 10:
            continue
        stt_val  = str(row[1]).strip()
        code_val = str(row[3]).strip()
        if not code_val or not stt_val:
            continue
        try:
            int(stt_val)
        except ValueError:
            continue  # skip header / non-data rows
        code     = code_val.upper()
        ton_dau  = _num(row[6])
        nhap_ky  = _num(row[7])
        xuat_ky  = _num(row[8])
        cuoi_raw = str(row[9]).strip()
        ton_cuoi = _num(cuoi_raw) if cuoi_raw and cuoi_raw not in ("-", "-   ", "  -   ") \
                   else (ton_dau + nhap_ky - xuat_ky)
        don_gia  = _num(row[10]) if len(row) > 10 else 0.0
        items[code] = {
            "stt":      stt_val,
            "kho":      str(row[2]).strip(),
            "code":     code_val,
            "name":     str(row[4]).strip(),
            "unit":     str(row[5]).strip(),
            "min_qty":  0,
            "ton_dau":  ton_dau,
            "nhap_ky":  nhap_ky,
            "xuat_ky":  xuat_ky,
            "ton_cuoi": ton_cuoi,
            "don_gia":  don_gia,
            "row":      i,
        }
    return items


def _num(val) -> float:
    """Parse numbers where ',' = thousands sep, '.' = decimal sep.
    Examples: '1,500' → 1500.0, '1,500.75' → 1500.75, '1.5' → 1.5
    """
    try:
        s = str(val).strip()
        if not s or s in ("None", ""):
            return 0.0
        s = s.replace(",", "")  # strip thousands commas
        return float(s)
    except (ValueError, TypeError):
        return 0.0


def update_stock(code: str, delta_nhap: float, delta_xuat: float):
    """
    Update columns H (Nhập kỳ), I (Xuất kỳ), J (Tồn cuối) in Bao_Cao_Ton_Kho for one item.
    Tồn cuối = Tồn đầu kỳ + Nhập kỳ - Xuất kỳ
    """
    ws = _ws()
    items = get_items()
    code_upper = code.strip().upper()
    if code_upper not in items:
        return
    item = items[code_upper]
    row   = item["row"]
    nhap  = item["nhap_ky"] + delta_nhap
    xuat  = item["xuat_ky"] + delta_xuat
    cuoi  = item["ton_dau"] + nhap - xuat
    ws.update_cell(row, 8,  nhap)   # H = Nhập kho
    ws.update_cell(row, 9,  xuat)   # I = Xuất bán
    ws.update_cell(row, 10, cuoi)   # J = Tồn cuối


def add_new_item(code: str, name: str, unit: str, qty: float,
                 kho: str = "", next_stt: int = None) -> dict:
    """Append a brand-new item row to Bao_Cao_Ton_Kho and return its data dict."""
    ws = _ws()
    if next_stt is None:
        items = get_items()
        existing_stts = [int(v["stt"]) for v in items.values() if str(v["stt"]).isdigit()]
        next_stt = max(existing_stts, default=0) + 1
    # A=empty, B=STT, C=Kho, D=Mã hàng, E=Tên, F=ĐVT, G=Tồn đầu, H=Nhập kỳ, I=Xuất kỳ, J=Tồn cuối
    ws.append_row(["", next_stt, kho, code, name, unit, 0, qty, 0, qty],
                  value_input_option="USER_ENTERED",
                  table_range="A1")
    logger.info("New item added to sheet: %s — %s", code, name)
    return {"stt": str(next_stt), "kho": kho, "code": code, "name": name,
            "unit": unit, "min_qty": 0, "ton_dau": 0,
            "nhap_ky": qty, "xuat_ky": 0, "ton_cuoi": qty}


def _ghiso_ws() -> gspread.Worksheet:
    spreadsheet = _get_gc().open_by_key(GOOGLE_SHEET_ID)
    try:
        return spreadsheet.worksheet("GHISO")
    except gspread.WorksheetNotFound:
        ws = spreadsheet.add_worksheet(title="GHISO", rows=5000, cols=12)
        ws.append_row(["Ngày", "Diễn giải / Nội dung", "Ngày XNI",
                        "Số LSX", "Số phiếu", "Loại phiếu", "STT",
                        "Mã NVL", "Tên Nguyên Vật Liệu", "DVT",
                        "Số Lượng", "Ghi Chú"])
        return ws


def append_transactions(doc_type: str, dien_giai: str, so_phieu: str,
                         so_lsx: str, doc_date: str, items_extracted: list[dict]):
    """Write all items to GHISO in a single batch call."""
    loai = "NK" if doc_type == "IN" else ("YC" if doc_type == "YC" else "XK")
    date_str = datetime.now().strftime("%m/%d/%Y")
    gh = _ghiso_ws()
    rows = [
        [
            date_str,
            dien_giai,
            doc_date,
            so_lsx,
            so_phieu,
            loai,
            i,
            item.get("code", ""),
            item.get("name", ""),
            item.get("unit", ""),
            float(item.get("quantity", 0)),
            item.get("note", ""),
        ]
        for i, item in enumerate(items_extracted, start=1)
    ]
    if rows:
        gh.append_rows(rows, value_input_option="USER_ENTERED")


# ── Claude extraction ──────────────────────────────────────────────────────────

SYSTEM_PROMPT = """You are a Vietnamese warehouse document parser for a manufacturing company.
Analyze the document and return ONLY a valid JSON:

{
  "doc_type": "IN" or "OUT" or "YC" or "UNKNOWN",
  "doc_ref": "số phiếu (e.g. NK250305/1 or XK250305/1 or DH250305/1)",
  "so_lsx": "số lệnh sản xuất if present, else empty string",
  "dien_giai": "diễn giải / nội dung (e.g. NKSX, NKM, XKSX, YCDKX)",
  "doc_date": "date in dd/mm/yyyy format",
  "items": [
    {
      "code": "mã hàng (item code)",
      "name": "tên hàng (item name)",
      "unit": "đơn vị tính",
      "quantity": numeric quantity,
      "unit_price": numeric unit price (đơn giá), 0 if not present,
      "note": "ghi chú if any"
    }
  ]
}

Rules:
- Return ONLY the JSON, no extra text.
- doc_type: "IN" for phiếu nhập kho (NK), "OUT" for phiếu xuất kho (XK), "YC" for phiếu dự kiến xuất theo đơn hàng (DK, DH, yêu cầu xuất).
- quantity must be a plain number.
- unit_price: extract from "đơn giá" / "giá" / "unit price" column if present, else 0.
- For phiếu xuất kho (OUT): use "số lượng thực nhận" as quantity; fall back to "số lượng".
- For phiếu dự kiến xuất (YC): use "số lượng" as quantity.
- For phiếu nhập kho (IN): use "số lượng" as quantity.
- Use "" for missing text fields, 0 for missing numbers.
- dien_giai examples: NKM (mua), NKSX (sản xuất nhập), XKSX (sản xuất xuất), YCDKX (dự kiến xuất).
"""


def _call_claude(content: list) -> dict:
    for attempt in range(3):
        try:
            resp = claude.messages.create(
                model="claude-sonnet-4-6",
                max_tokens=8192,
                system=SYSTEM_PROMPT,
                messages=[{"role": "user", "content": content}],
            )
            raw = resp.content[0].text.strip()
            # Strip markdown code fences if present
            if raw.startswith("```"):
                raw = re.sub(r"^```[a-z]*\n?", "", raw)
                raw = re.sub(r"\n?```$", "", raw)
                raw = raw.strip()
            try:
                return json.loads(raw)
            except json.JSONDecodeError:
                s, e = raw.find("{"), raw.rfind("}") + 1
                if s != -1 and e > s:
                    try:
                        return json.loads(raw[s:e])
                    except json.JSONDecodeError:
                        logger.warning("Claude JSON parse failed on attempt %d", attempt + 1)
                if attempt < 2:
                    time.sleep(5)
                    continue
                return {}
        except anthropic.APIStatusError as e:
            if e.status_code == 529 and attempt < 2:
                wait = 10 * (attempt + 1)
                logger.warning("Claude overloaded, retry in %ds", wait)
                time.sleep(wait)
            else:
                raise
    return {}


def extract_from_pdf(pdf_bytes: bytes, file_name: str) -> dict:
    pdf_data = base64.standard_b64encode(pdf_bytes).decode("utf-8")
    return _call_claude([
        {"type": "document", "source": {"type": "base64", "media_type": "application/pdf", "data": pdf_data}},
        {"type": "text", "text": f'Parse this warehouse document: "{file_name}"'},
    ])


def extract_from_image(image_bytes: bytes, mime_type: str) -> dict:
    img_data = base64.standard_b64encode(image_bytes).decode("utf-8")
    return _call_claude([
        {"type": "image", "source": {"type": "base64", "media_type": mime_type, "data": img_data}},
        {"type": "text", "text": "Parse this warehouse document (phiếu nhập/xuất kho)."},
    ])


def extract_from_excel(excel_bytes: bytes) -> dict:
    """
    Parse PNK/PXK Excel files directly with openpyxl.
    Detects doc type, date, số phiếu, and item rows automatically.
    Falls back to Claude (text prompt) if structure is unrecognised.
    """
    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), read_only=True, data_only=True)

    # ── 1. Pick the most relevant sheet ──────────────────────────────────────
    sheet_priority = ["PNK", "PXK", "NHAP", "XUAT", "XK", "NK", "PHIEU"]
    ws = None
    for name in sheet_priority:
        for sn in wb.sheetnames:
            if name in sn.upper():
                ws = wb[sn]
                break
        if ws:
            break
    if ws is None:
        ws = wb.active   # fallback: first sheet

    rows = [list(r) for r in ws.iter_rows(values_only=True)]

    # ── 2. Detect doc type from sheet name or cell content ───────────────────
    sheet_name_upper = ws.title.upper()
    full_text = " ".join(str(c) for r in rows[:20] for c in r if c)

    DK_KEYWORDS = ("DỰ KIẾN XUẤT", "DU KIEN XUAT", "PHIẾU DỰ KIẾN", "PHIEU DU KIEN",
                   "XUẤT THEO ĐƠN HÀNG", "XUAT THEO DON HANG")
    if "NHẬP KHO" in full_text or "PNK" in sheet_name_upper or "NHAP" in sheet_name_upper:
        doc_type = "IN"
        dien_giai = "NKM"
    elif any(k in full_text.upper() for k in DK_KEYWORDS):
        doc_type = "YC"
        dien_giai = "YCDKX"
    elif "XUẤT KHO" in full_text or "PXK" in sheet_name_upper or "XUAT" in sheet_name_upper:
        doc_type = "OUT"
        dien_giai = "XKSX"
    else:
        doc_type = "UNKNOWN"
        dien_giai = ""

    # ── 3. Extract header fields (scan first 15 rows) ────────────────────────
    doc_date = ""
    so_phieu = ""
    so_lsx   = ""

    date_keywords  = ["ngày nhập", "ngày xuất", "ngày", "date"]
    phieu_keywords = ["số phiếu", "so phieu", "phiếu số"]
    lsx_keywords   = ["số lsx", "lệnh sx", "lsx"]

    for row in rows[:15]:
        row_text = [str(c).strip().lower() if c else "" for c in row]
        full_row = " ".join(row_text)
        # Find value in next non-empty cell after label
        for i, cell in enumerate(row_text):
            if any(k in cell for k in date_keywords) and not doc_date:
                val = next((row[j] for j in range(i+1, len(row)) if row[j]), None)
                if isinstance(val, datetime):
                    doc_date = val.strftime("%m/%d/%Y")
                elif val:
                    doc_date = str(val).strip()
            if any(k in cell for k in phieu_keywords) and not so_phieu:
                val = next((row[j] for j in range(i+1, len(row)) if row[j]), None)
                if val:
                    so_phieu = str(val).strip()
            if any(k in cell for k in lsx_keywords) and not so_lsx:
                val = next((row[j] for j in range(i+1, len(row)) if row[j]), None)
                if val:
                    so_lsx = str(val).strip()

    # ── 4. Find item table header row ────────────────────────────────────────
    # Look for a row that has "mã hàng" or "stt" to find the column positions
    col_stt  = col_code = col_name = col_unit = col_qty = col_note = col_ton_sau_dk = col_price = None
    header_row_idx = None

    CODE_KEYS    = ("mã hàng", "ma hang", "mã nvl", "ma nvl", "mã vt", "ma vt", "item code")
    NAME_KEYS    = ("tên hàng", "ten hang", "tên nguyên vật liệu", "ten nguyen vat lieu", "tên nvl", "ten nvl", "item name")
    UNIT_KEYS    = ("đvt", "dvt", "đơn vị", "don vi", "unit")
    QTY_REAL     = ("thực nhận", "thuc nhan", "sl thực nhận", "sl thuc nhan")
    QTY_KEYS     = ("số lượng", "so luong", "sl", "qty", "quantity")
    NOTE_KEYS    = ("ghi chú", "ghi chu", "note", "ghi chú/note")
    TON_SAU_KEYS = ("tồn kho sau dự kiến", "ton kho sau du kien", "sau dự kiến", "sau dk", "tồn sau dk")
    PRICE_KEYS   = ("đơn giá", "don gia", "giá bán", "gia ban", "đg", "unit price", "price")

    for i, row in enumerate(rows):
        row_lower = [str(c).strip().lower() if c else "" for c in row]
        if any(any(k in c for k in CODE_KEYS) for c in row_lower):
            header_row_idx = i
            # Scan this row for column positions
            for j, cell in enumerate(row_lower):
                if "stt" in cell and col_stt is None:                              col_stt        = j
                if any(k in cell for k in CODE_KEYS) and col_code is None:        col_code       = j
                if any(k in cell for k in NAME_KEYS) and col_name is None:        col_name       = j
                if any(k in cell for k in UNIT_KEYS) and col_unit is None:        col_unit       = j
                if any(k in cell for k in QTY_REAL):                              col_qty        = j
                elif any(k in cell for k in QTY_KEYS) and col_qty is None:        col_qty        = j
                if any(k in cell for k in NOTE_KEYS) and col_note is None:        col_note       = j
                if any(k in cell for k in TON_SAU_KEYS) and col_ton_sau_dk is None: col_ton_sau_dk = j
                if any(k in cell for k in PRICE_KEYS) and col_price is None:      col_price      = j
            # Also scan next 1-2 rows for sub-headers (e.g. "thực nhận" on second header row)
            for sub_row in rows[i + 1: i + 3]:
                sub_lower = [str(c).strip().lower() if c else "" for c in sub_row]
                for j, cell in enumerate(sub_lower):
                    if any(k in cell for k in QTY_REAL):
                        col_qty = j
                        header_row_idx = i + (rows[i + 1: i + 3].index(sub_row) + 1)
                    if any(k in cell for k in UNIT_KEYS) and col_unit is None:        col_unit       = j
                    if any(k in cell for k in NOTE_KEYS) and col_note is None:        col_note       = j
                    if any(k in cell for k in TON_SAU_KEYS) and col_ton_sau_dk is None: col_ton_sau_dk = j
                    if any(k in cell for k in PRICE_KEYS) and col_price is None:      col_price      = j
            break

    # ── 5. Extract item rows ──────────────────────────────────────────────────
    logger.info("Excel parse: sheet=%s header_row=%s col_code=%s col_qty=%s col_name=%s col_unit=%s",
                ws.title, header_row_idx, col_code, col_qty, col_name, col_unit)
    items = []
    if header_row_idx is not None and col_qty is not None:
        for row in rows[header_row_idx + 1:]:
            if not row or all(c is None for c in row):
                continue
            qty_val = row[col_qty] if col_qty < len(row) else None
            qty_str = str(qty_val).replace(",", "").strip() if qty_val is not None else ""
            try:
                qty = float(qty_str)
            except (ValueError, TypeError):
                qty = 0.0
            if qty == 0:
                logger.debug("Skipped row (qty=0 or dash): %s", [str(c)[:20] for c in row[:6]])
                continue

            code = str(row[col_code]).strip() if col_code is not None and col_code < len(row) and row[col_code] else ""
            name = str(row[col_name]).strip() if col_name is not None and col_name < len(row) and row[col_name] else ""
            unit = str(row[col_unit]).strip() if col_unit is not None and col_unit < len(row) and row[col_unit] else ""
            note = str(row[col_note]).strip() if col_note is not None and col_note < len(row) and row[col_note] else ""

            if not code and not name:
                continue

            ton_sau_dk = _num(row[col_ton_sau_dk]) if col_ton_sau_dk is not None and col_ton_sau_dk < len(row) and row[col_ton_sau_dk] else None
            unit_price = _num(row[col_price]) if col_price is not None and col_price < len(row) and row[col_price] else 0.0
            items.append({"code": code, "name": name, "unit": unit, "quantity": qty, "note": note,
                          "ton_kho_sau_dk": ton_sau_dk, "unit_price": unit_price})

    logger.info("Excel parse: found %d items", len(items))

    # ── 6. Fallback: send text to Claude if items not found ──────────────────
    if not items or doc_type == "UNKNOWN":
        logger.warning("Excel auto-parse incomplete — falling back to Claude text extraction")
        # Build plain text summary of the sheet for Claude
        text_rows = []
        for row in rows:
            line = "\t".join(str(c) if c is not None else "" for c in row)
            if line.strip():
                text_rows.append(line)
        sheet_text = "\n".join(text_rows)
        return _call_claude([
            {"type": "text", "text": f"Parse this warehouse Excel sheet data:\n\n{sheet_text}"},
        ])

    return {
        "doc_type":   doc_type,
        "doc_ref":    so_phieu,
        "so_lsx":     so_lsx,
        "dien_giai":  dien_giai,
        "doc_date":   doc_date,
        "items":      items,
    }


# ── Format reply ───────────────────────────────────────────────────────────────

def format_result(data: dict, items_after: dict) -> str:
    doc_type = data.get("doc_type", "UNKNOWN")
    if doc_type not in ("IN", "OUT", "YC"):
        return "⚠️ Không xác định được loại chứng từ (nhập/xuất kho/dự kiến xuất)."

    icon      = "📥" if doc_type == "IN" else ("📋" if doc_type == "YC" else "📤")
    label     = "NHẬP KHO" if doc_type == "IN" else ("DỰ KIẾN XUẤT" if doc_type == "YC" else "XUẤT KHO")
    so_phieu  = data.get("doc_ref", "—")
    ngay      = data.get("doc_date", "—")
    dien_giai = data.get("dien_giai", "")

    lines = [
        f"{icon} *{label}*",
        f"📅 Ngày: {ngay}   |   Số phiếu: {so_phieu}",
    ]
    if dien_giai:
        lines.append(f"📝 Diễn giải: {dien_giai}")
    lines.append("")
    lines.append("*Chi tiết hàng hoá:*")

    warnings = []
    for item in data.get("items", []):
        code  = str(item.get("code", "")).strip().upper()
        name  = item.get("name", code)
        unit  = item.get("unit", "")
        qty   = item.get("quantity", 0)
        # Read Tồn cuối (col J) from updated DMHANGHOA
        row_data = items_after.get(code, {})
        ton_cuoi = row_data.get("ton_cuoi", 0)
        min_qty  = row_data.get("min_qty", 0)

        arrow = "▲" if doc_type == "IN" else "▼"
        lines.append(f"  • *{name}* `{code}`")
        lines.append(f"    {arrow} {qty} {unit}  →  Tồn cuối: *{ton_cuoi:.0f}* {unit}")

        warn = stock_warning(ton_cuoi, min_qty)
        if warn:
            warnings.append(f"  `{code}` — {warn}")

    lines.append("\n✅ Đã cập nhật Tồn cuối (cột J)")

    if warnings:
        lines.append("\n⚠️ *Cảnh báo tồn kho:*")
        lines.extend(warnings)

    return "\n".join(lines)


def _get_min_map() -> dict[str, float]:
    """Read Tồn mức MIN from DMHANGHOA sheet (col F, index 5)."""
    try:
        dm_ws = _get_gc().open_by_key(GOOGLE_SHEET_ID).worksheet("DMHANGHOA")
        rows = dm_ws.get_all_values()
        result = {}
        for row in rows[1:]:
            if len(row) > 5 and row[2]:
                result[row[2].strip().upper()] = _num(row[5])
        return result
    except Exception as e:
        logger.warning("Could not read MIN from DMHANGHOA: %s", e)
        return {}


def apply_stock_colors(items_extracted: list[dict], doc_type: str,
                        items: dict = None) -> list[dict]:
    """
    Before writing to GHISO, compare each item's quantity against TỒN CUỐI
    and color columns R/S/T in Bao_Cao_Ton_Kho.

    ratio = quantity / TỒN CUỐI
    Col R = Yellow  50% ≤ ratio < 70%
    Col S = Orange  70% ≤ ratio ≤ 90%
    Col T = Red     ratio > 90%

    Returns list of dicts with color, code, name, unit, qty, ton_cuoi, remaining.
    """
    ws = _ws()
    if items is None:
        items = get_items()
    results = []
    fmt_requests = []

    for extracted in items_extracted:
        code = str(extracted.get("code", "")).strip().upper()
        if not code or code not in items:
            continue
        item = items[code]
        if "row" not in item:
            continue
        row = item["row"]
        ton_cuoi = item["ton_cuoi"]
        qty      = float(extracted.get("quantity", 0))
        name     = item.get("name", extracted.get("name", code))
        unit     = item.get("unit", extracted.get("unit", ""))

        if ton_cuoi <= 0 or qty <= 0:
            continue

        ratio = qty / ton_cuoi

        # For YC: use "TỒN KHO SAU DỰ KIẾN" from the Excel document (col N)
        if doc_type == "YC":
            ton_sau = extracted.get("ton_kho_sau_dk")
            remaining = ton_sau if ton_sau is not None else (ton_cuoi - qty)
        elif doc_type == "OUT":
            remaining = ton_cuoi - qty
        else:
            remaining = ton_cuoi + qty

        if ratio > 0.90:
            color = "🔴 Đỏ"
            fmt_requests.append({"range": f"T{row}", "format": {"backgroundColor": {"red": 1.0, "green": 0.0, "blue": 0.0}}})
        elif ratio >= 0.70:
            color = "🟠 Cam"
            fmt_requests.append({"range": f"S{row}", "format": {"backgroundColor": {"red": 1.0, "green": 0.5, "blue": 0.0}}})
        elif ratio >= 0.50:
            color = "🟡 Vàng"
            fmt_requests.append({"range": f"R{row}", "format": {"backgroundColor": {"red": 1.0, "green": 1.0, "blue": 0.0}}})
        else:
            continue

        results.append({
            "color":     color,
            "code":      code,
            "name":      name,
            "unit":      unit,
            "qty":       qty,
            "ton_cuoi":  ton_cuoi,
            "remaining": remaining,
            "ratio":     ratio,
        })

    if fmt_requests:
        ws.batch_format(fmt_requests)

    return results


# ── Price resolution & invoice ────────────────────────────────────────────────

def _get_bang_gia_prices() -> dict[str, float]:
    """Read Mã hàng (col D, index 3) and Đơn giá 1 (col G, index 6) from BANG_GIA sheet."""
    try:
        ws = _get_gc().open_by_key(GOOGLE_SHEET_ID).worksheet("BANG_GIA")
        rows = ws.get_all_values()
        prices = {}
        for row in rows:
            if len(row) < 7:
                continue
            code = str(row[3]).strip().upper()
            if not code:
                continue
            price = _num(row[6])
            if price > 0:
                prices[code] = price
        return prices
    except Exception as e:
        logger.warning("Could not read BANG_GIA prices: %s", e)
        return {}


def _resolve_prices(items_extracted: list[dict]) -> list[dict]:
    """Fill unit_price from BANG_GIA sheet (col G) when not present in document."""
    price_map = _get_bang_gia_prices()
    result = []
    for item in items_extracted:
        code      = str(item.get("code", "")).strip().upper()
        doc_price = float(item.get("unit_price", 0) or 0)
        price     = doc_price if doc_price > 0 else price_map.get(code, 0)
        result.append({**item, "unit_price": price})
    return result


def create_invoice_sheet(doc_ref: str, dien_giai: str, doc_date: str,
                          so_lsx: str, items_with_price: list[dict]) -> tuple[str, str]:
    """Create a formatted invoice tab. Returns (sheet_name, sheet_url)."""
    spreadsheet = _get_gc().open_by_key(GOOGLE_SHEET_ID)

    # Unique sheet name
    base = f"HĐ_{doc_ref}" if doc_ref else f"HĐ_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    base = re.sub(r"[\\/*?:\[\]]", "_", base)[:50]
    name = base
    existing = {ws.title for ws in spreadsheet.worksheets()}
    suffix = 2
    while name in existing:
        name = f"{base}_{suffix}"
        suffix += 1

    n_items = len(items_with_price)
    ws = spreadsheet.add_worksheet(title=name, rows=n_items + 12, cols=7)

    grand_total = sum(
        float(i.get("quantity", 0)) * float(i.get("unit_price", 0))
        for i in items_with_price
    )

    # Build rows
    rows_data = [
        ["PHIẾU XUẤT KHO", "", "", "", "", "", ""],
        ["Số phiếu:", doc_ref, "", "", "Ngày:", doc_date, ""],
        ["Diễn giải:", dien_giai, "", "", "Số LSX:", so_lsx, ""],
        ["", "", "", "", "", "", ""],
        ["STT", "Mã hàng", "Tên hàng", "ĐVT", "Số lượng", "Đơn giá", "Thành tiền"],
    ]
    for idx, item in enumerate(items_with_price, start=1):
        qty   = float(item.get("quantity", 0))
        price = float(item.get("unit_price", 0))
        rows_data.append([
            idx,
            item.get("code", ""),
            item.get("name", ""),
            item.get("unit", ""),
            qty,
            price,
            qty * price,
        ])
    rows_data.append(["TỔNG CỘNG", "", "", "", "", "", grand_total])

    ws.update(rows_data, "A1", value_input_option="USER_ENTERED")

    # Formatting via Sheets API
    sid = ws.id
    data_end = 5 + n_items  # 0-based last data row index (inclusive)
    spreadsheet.batch_update({"requests": [
        # Merge title A1:G1
        {"mergeCells": {
            "range": {"sheetId": sid, "startRowIndex": 0, "endRowIndex": 1,
                      "startColumnIndex": 0, "endColumnIndex": 7},
            "mergeType": "MERGE_ALL",
        }},
        # Title style
        {"repeatCell": {
            "range": {"sheetId": sid, "startRowIndex": 0, "endRowIndex": 1,
                      "startColumnIndex": 0, "endColumnIndex": 7},
            "cell": {"userEnteredFormat": {
                "textFormat": {"bold": True, "fontSize": 14},
                "horizontalAlignment": "CENTER",
            }},
            "fields": "userEnteredFormat(textFormat,horizontalAlignment)",
        }},
        # Header row (row 5, index 4) — bold + light blue bg
        {"repeatCell": {
            "range": {"sheetId": sid, "startRowIndex": 4, "endRowIndex": 5,
                      "startColumnIndex": 0, "endColumnIndex": 7},
            "cell": {"userEnteredFormat": {
                "textFormat": {"bold": True},
                "horizontalAlignment": "CENTER",
                "backgroundColor": {"red": 0.82, "green": 0.91, "blue": 0.98},
            }},
            "fields": "userEnteredFormat(textFormat,horizontalAlignment,backgroundColor)",
        }},
        # Total row — bold
        {"repeatCell": {
            "range": {"sheetId": sid, "startRowIndex": data_end, "endRowIndex": data_end + 1,
                      "startColumnIndex": 0, "endColumnIndex": 7},
            "cell": {"userEnteredFormat": {"textFormat": {"bold": True}}},
            "fields": "userEnteredFormat(textFormat)",
        }},
        # Number format for Đơn giá (col F=5) and Thành tiền (col G=6)
        {"repeatCell": {
            "range": {"sheetId": sid, "startRowIndex": 5, "endRowIndex": data_end + 1,
                      "startColumnIndex": 5, "endColumnIndex": 7},
            "cell": {"userEnteredFormat": {
                "numberFormat": {"type": "NUMBER", "pattern": "#,##0"},
            }},
            "fields": "userEnteredFormat(numberFormat)",
        }},
    ]})

    sheet_url = (
        f"https://docs.google.com/spreadsheets/d/{GOOGLE_SHEET_ID}"
        f"/edit#gid={sid}"
    )
    return name, sheet_url


# ── Process document ───────────────────────────────────────────────────────────

async def _process(update: Update, data: dict, file_name: str):
    doc_type = data.get("doc_type", "UNKNOWN")
    if doc_type not in ("IN", "OUT", "YC") or not data.get("items"):
        await update.message.reply_text(
            f"⚠️ Không đọc được chứng từ từ *{file_name}*.\n"
            "Hãy đảm bảo đây là phiếu nhập kho hoặc phiếu xuất kho.",
            parse_mode="Markdown",
        )
        return

    # Single sheet read for the entire processing flow
    existing = get_items()

    # Auto-add new items to Bao_Cao_Ton_Kho for IN documents
    new_items_added = []
    if doc_type == "IN":
        next_stt = max(
            (int(v["stt"]) for v in existing.values() if str(v["stt"]).isdigit()),
            default=0
        ) + 1
        for item in data.get("items", []):
            code = str(item.get("code", "")).strip().upper()
            if code and code not in existing:
                qty = float(item.get("quantity", 0))
                new_item = add_new_item(code, item.get("name", ""), item.get("unit", ""), qty,
                                        next_stt=next_stt)
                new_items_added.append(item)
                existing[code] = new_item  # cache so duplicates and format_result can use it
                next_stt += 1

    # Resolve unit prices and create invoice for OUT documents
    invoice_name = invoice_url = None
    if doc_type == "OUT":
        items_priced = _resolve_prices(data.get("items", []))
        data = {**data, "items": items_priced}
        invoice_name, invoice_url = create_invoice_sheet(
            doc_ref          = data.get("doc_ref", ""),
            dien_giai        = data.get("dien_giai", ""),
            doc_date         = data.get("doc_date", ""),
            so_lsx           = data.get("so_lsx", ""),
            items_with_price = items_priced,
        )

    # Color Bao_Cao_Ton_Kho BEFORE writing to GHISO (reuse existing — no extra sheet read)
    color_results = apply_stock_colors(data.get("items", []), doc_type, items=existing)

    # Write to GHISO
    append_transactions(
        doc_type        = doc_type,
        dien_giai       = data.get("dien_giai", ""),
        so_phieu        = data.get("doc_ref", ""),
        so_lsx          = data.get("so_lsx", ""),
        doc_date        = data.get("doc_date", ""),
        items_extracted = data.get("items", []),
    )

    # Reply with document summary (reuse existing — no extra sheet read)
    reply = format_result(data, existing)
    while reply:
        chunk, reply = reply[:4096], reply[4096:]
        try:
            await update.message.reply_text(chunk, parse_mode="Markdown")
        except Exception:
            await update.message.reply_text(chunk)

    # Send detailed color warning after GHISO write
    if color_results:
        lines = ["⚠️ CẢNH BÁO TỒN KHO:\n"]
        for r in color_results:
            lines.append(
                f"{r['color']}  {r['name']} [{r['code']}]\n"
                f"  Xuất: {r['qty']:.0f} {r['unit']} | "
                f"Tồn trước: {r['ton_cuoi']:.0f} | "
                f"Còn lại: {r['remaining']:.0f} {r['unit']} "
                f"({r['ratio']*100:.0f}%)"
            )
        msg = "\n".join(lines)
        while len(msg) > 4096:
            split = msg.rfind("\n", 0, 4096)
            await update.message.reply_text(msg[:split])
            msg = msg[split:].lstrip("\n")
        await update.message.reply_text(msg)

    # Invoice notification for OUT documents
    if invoice_name and invoice_url:
        lines = ["🧾 *Hoá đơn xuất kho đã được tạo:*\n"]
        lines.append(f"📄 Sheet: `{invoice_name}`")
        items_priced = data.get("items", [])
        grand_total  = sum(float(i.get("quantity", 0)) * float(i.get("unit_price", 0))
                           for i in items_priced)
        for item in items_priced:
            qty   = float(item.get("quantity", 0))
            price = float(item.get("unit_price", 0))
            total = qty * price
            src   = "" if price > 0 else " _(chưa có đơn giá)_"
            lines.append(
                f"  • *{item.get('name', '')}* `{str(item.get('code','')).upper()}`\n"
                f"    {qty:.0f} {item.get('unit','')} × {price:,.0f} = *{total:,.0f}*{src}"
            )
        lines.append(f"\n💰 *Tổng cộng: {grand_total:,.0f}*")
        lines.append(f"\n🔗 [Mở hoá đơn]({invoice_url})")
        await update.message.reply_text("\n".join(lines), parse_mode="Markdown")

    # Notify about newly added items
    if new_items_added:
        lines = [
            "✨ *Hàng hoá mới đã được thêm vào danh mục:*\n",
            f"📋 Phiếu: {data.get('doc_ref', '—')}  |  Ngày: {data.get('doc_date', '—')}\n",
        ]
        for item in new_items_added:
            code = str(item.get("code", "")).strip().upper()
            name = item.get("name", "")
            unit = item.get("unit", "")
            qty  = float(item.get("quantity", 0))
            lines.append(
                f"  • *{name}* `{code}`\n"
                f"    ĐVT: {unit}  |  Tồn ban đầu: *{qty:.0f}* {unit}"
            )
        lines.append("\n📌 _Nhớ cập nhật Tồn mức MIN cho hàng mới trong sheet._")
        await update.message.reply_text("\n".join(lines), parse_mode="Markdown")


# ── Telegram handlers ──────────────────────────────────────────────────────────

def _allowed(update: Update) -> bool:
    if TARGET_GROUP_ID and str(update.effective_chat.id) != str(TARGET_GROUP_ID):
        return False
    return True


async def handle_document(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not _allowed(update):
        return
    doc = update.message.document
    if not doc:
        return

    mime      = doc.mime_type or ""
    file_name = doc.file_name or "document"
    ext       = os.path.splitext(file_name)[1].lower()

    is_pdf   = mime == "application/pdf"
    is_image = mime.startswith("image/")
    is_excel = ext in (".xlsx", ".xls", ".xlsm", ".xlsb") or mime in (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel",
        "application/vnd.ms-excel.sheet.macroEnabled.12",
    )

    if not is_pdf and not is_image and not is_excel:
        return

    processing = await update.message.reply_text(
        f"⏳ Đang đọc *{file_name}*...", parse_mode="Markdown"
    )
    try:
        tg_file = await context.bot.get_file(doc.file_id)
        buf = io.BytesIO()
        await tg_file.download_to_memory(buf)
        file_bytes = buf.getvalue()

        if is_pdf:
            data = extract_from_pdf(file_bytes, file_name)
        elif is_excel:
            data = extract_from_excel(file_bytes)
        else:
            data = extract_from_image(file_bytes, mime)

        await processing.delete()
        await _process(update, data, file_name)
    except Exception as e:
        logger.error("handle_document error: %s", e, exc_info=True)
        await update.message.reply_text(f"❌ Lỗi xử lý {file_name}: {e}")


async def handle_photo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not _allowed(update):
        return

    photo = update.message.photo[-1]
    processing = await update.message.reply_text("⏳ Đang đọc ảnh chứng từ...")
    try:
        tg_file = await context.bot.get_file(photo.file_id)
        buf = io.BytesIO()
        await tg_file.download_to_memory(buf)

        data = extract_from_image(buf.getvalue(), "image/jpeg")
        await processing.delete()
        await _process(update, data, "ảnh chứng từ")
    except Exception as e:
        logger.error("handle_photo error: %s", e, exc_info=True)
        await update.message.reply_text(f"❌ Lỗi xử lý ảnh: {e}")


async def stock_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """/stock [mã hàng] — show Tồn cuối (col J). No arg = warnings only."""
    try:
        items = get_items()   # reads Tồn cuối from col J
        query = " ".join(context.args).strip().upper() if context.args else ""

        if query:
            matched = {
                code: item for code, item in items.items()
                if query in code or query in item["name"].upper()
            }
            if not matched:
                await update.message.reply_text(f"❌ Không tìm thấy mã/tên hàng: `{query}`", parse_mode="Markdown")
                return

            lines = [f"📦 *Tồn kho — {query}*\n"]
            for code, item in matched.items():
                qty  = item["ton_cuoi"]
                warn = stock_warning(qty, item["min_qty"])
                flag = f"\n    {warn}" if warn else ""
                lines.append(
                    f"• *{item['name']}* (`{code}`)\n"
                    f"  Kho: {item['kho']} | Tồn cuối: *{qty:.0f}* {item['unit']} | MIN: {item['min_qty']:.0f}{flag}"
                )
        else:
            lines = ["📦 *Cảnh báo tồn kho*\n"]
            warned = []
            for code, item in items.items():
                qty  = item["ton_cuoi"]
                warn = stock_warning(qty, item["min_qty"])
                if warn:
                    warned.append((code, item, qty, warn))

            if not warned:
                await update.message.reply_text("✅ Tất cả hàng hoá đều trên mức cảnh báo.")
                return

            level_order = {"🔴": 0, "🟠": 1, "🟢": 2}
            warned.sort(key=lambda x: level_order.get(x[3][0], 9))

            for code, item, qty, warn in warned:
                lines.append(
                    f"• *{item['name']}* (`{code}`) — Kho: {item['kho']}\n"
                    f"  {warn}"
                )

        msg = "\n".join(lines)
        # Split if > 4096 chars
        while len(msg) > 4096:
            split = msg.rfind("\n", 0, 4096)
            await update.message.reply_text(msg[:split], parse_mode="Markdown")
            msg = msg[split:].lstrip("\n")
        await update.message.reply_text(msg, parse_mode="Markdown")

    except Exception as e:
        logger.error("stock_handler error: %s", e, exc_info=True)
        await update.message.reply_text(f"❌ Lỗi: {e}")


async def start_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "🏭 *Warehouse Bot*\n\n"
        "Gửi *PDF* hoặc *ảnh* phiếu nhập/xuất kho để cập nhật tồn kho.\n\n"
        "Lệnh:\n"
        "  /stock — Xem cảnh báo tồn kho\n"
        "  /stock \\[mã hàng\\] — Tìm kiếm theo mã hoặc tên hàng\n\n"
        "Mức cảnh báo tồn kho:\n"
        "  🟢 Mức 1 — ≤ 50% tồn MIN\n"
        "  🟠 Mức 2 — ≤ 30% tồn MIN\n"
        "  🔴 Mức 3 — ≤ 20% tồn MIN — Cần nhập ngay",
        parse_mode="Markdown",
    )


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    logger.info("Initialising warehouse sheets...")
    ensure_sheets()

    app = Application.builder().token(TELEGRAM_BOT_TOKEN).build()
    app.add_handler(CommandHandler("start", start_handler))
    app.add_handler(CommandHandler("stock", stock_handler))
    app.add_handler(MessageHandler(filters.Document.ALL, handle_document))
    app.add_handler(MessageHandler(filters.PHOTO, handle_photo))

    logger.info("Warehouse Bot running.")
    app.run_polling()


if __name__ == "__main__":
    main()
